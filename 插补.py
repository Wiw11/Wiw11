import openpyxl as xl

workBook = xl.load_workbook('MK值.xlsx', data_only=True)
sheet1 = workBook['全年准确']

temp = dict()
for i in range(3, 236):
    temp[sheet1.cell(row=i, column=1).value] = (sheet1.cell(row=i, column=4).value,sheet1.cell(row=i, column=5).value)
# print(temp)

sheet2 = workBook['插补表']
id = [54766,54765,57086,57087,57195,57198]
# for i in range(2,77):
#     id.append(sheet2.cell(row = i,column = 1).value)

s = dict()
for item in id:
    m = 100
    ls = []
    for i in temp.keys():
        if i == item:
            continue
        else:
            n = ((temp.get(item)[0] - temp.get(i)[0])**2 + (temp.get(item)[1] - temp.get(i)[1])**2)**0.5
            ls.append(n)
            if 0.2 < n <= m:
                s[item] = i
                m = n
            else:
                continue
    # ls.sort()
    # print(ls[0])
print(s)
# f = open('插补表.txt','w')
# f.write('插补站    临近站\n')
# for i in s.items():
#     f.write(str(i[0])+'    '+str(i[1]))
#     f.write('\n')
# f.close()