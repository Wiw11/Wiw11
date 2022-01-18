import openpyxl as xl
import os
import re

wb = input('请输入工作簿名称：')
# file_name = input('请输入新建文件夹名称：')
workBook = xl.load_workbook(wb+'.xlsx', data_only=True)
station = workBook.get_sheet_names()
# os.mkdir(wb)  #创建写入文件夹

s = []
for lable in station:
    sheet = workBook[lable]
    # avtemp = []  # 平均气温
    # for i in range(3, 65):
    #     avtemp.append(sheet.cell(row=i, column=6).value)
##    name = ''.join(re.findall('[\u4e00-\u9fa5]', lable))
    # m = re.findall('(\d{5})(?=.*)',lable.replace(' ',''))
    # name = lable.replace(m[0],'').replace(' ','')
    name = lable.replace(' ','')

    # if len(name) != 0:
    #     pass
    # else:
    #     name = str(sheet.cell(row=1, column=16).value)

    # data = []
    # for item in avtemp:
    #     if item == None or item == 0:
    #         break
    #     else:
    #         data.append(str(round(item,1)) + "\n")
    #
    # y = []  #年份
    # for i in range(3,len(data)+2):
    #     y.append(sheet.cell(row=i, column=1).value)

    if sheet.cell(row=61, column=1).value != 2020:
        s.append(name + '\n')

    # if y[-1]-y[0]+1 == len(y):
    #     f = open('D:\\课程文件\\气候变化\\'+ file_name + '\\' + name + ".txt", 'w')
    # else:
    #     f = open('D:\\课程文件\\气候变化\\'+ file_name + '\\' + '#'+ name + ".txt", 'w')


    # data = []
    # for item in avtemp:
    #     data.append(str(round(item,1)) + "\n")
    #
    # f = open('D:\\课程文件\\气候变化\\' + wb + '\\' + name + ".txt", 'w')
    # f.write(name + '\n')
    # f.write(wb[:2] + '平均气温'+ '\n')
    # f.write('1960\n')
    # f.write('2020\n')
    # f.writelines(data)

f = open('冬季插补.txt', 'w')
f.writelines(s)
f.close()
