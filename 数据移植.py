import openpyxl as xl

workBook1 = xl.load_workbook('平均气温 全年.xlsx', data_only=True)
station1 = workBook1.get_sheet_names()

workBook2 = xl.load_workbook('冬季气温.xlsx')
for lable in station1:
    sheet1 = workBook1[lable]
    sheet2 = workBook2.create_sheet(lable)
    n = 10
    for i in range(1,64):
        sheet2.cell(row=i, column=1,value = sheet1.cell(row=i, column=1).value)
        sheet2.cell(row=i, column=2,value = sheet1.cell(row=i, column=n+2).value)
        sheet2.cell(row=i, column=3,value = sheet1.cell(row=i, column=n+3).value)
        sheet2.cell(row=i, column=4,value = sheet1.cell(row=i, column=n-8).value)
workBook2.save('冬季气温.xlsx')