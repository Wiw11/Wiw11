import openpyxl as xl

# file_name = input('请输入文件名称：')
workBook1 = xl.load_workbook('D:\\课程文件\\气候变化\\AA113400.xlsx', data_only=True)
station = workBook1.get_sheet_names()
print(len(station))
workBook2 = xl.load_workbook('D:\\课程文件\\气候变化\\分站数据整合.xlsx', data_only=True)
for name in station:
    try:
        sheet1 = workBook1[name]
        name2 = name.replace(' ','')
        sheet2 = workBook2[name2]
        for i in range(2,63):
            sheet2.cell(row=i+1, column=3).value = sheet1.cell(row=i, column=6).value
    except:
        print(name)
workBook2.save('D:\\课程文件\\气候变化\\分站数据整合.xlsx')
